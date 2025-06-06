import streamlit as st
import pandas as pd
import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ページ設定
st.set_page_config(page_title="Excel差分ツール", page_icon="🌿", layout="centered")

# カスタムCSS（ロゴ非表示、ファイルアップロード日本語化トライアル）
st.markdown("""
<style>
    body {
        background-color: #e6f4e6;
    }
    .main {
        background-color: #ffffff;
        border-radius: 10px;
        padding: 2rem;
    }
    .upload-label {
        font-size: 1rem;
        font-weight: bold;
        margin-bottom: 0.2rem;
        display: block;
    }
    .stFileUploader {
        margin-bottom: 2rem;
    }
</style>
""", unsafe_allow_html=True)

# タイトルと説明
st.title("Excel差分比較ツール")
st.markdown("""
#### 📜 使い方：
1. 「暫定データファイル」と「確定データファイル」を選んでください。
2. 差分を自動で抽出して表示します。
3. Excelファイルとしてダウンロードも可能です。
""")

# アップローダー
st.markdown('<label class="upload-label">📂 暫定データファイルをアップロード</label>', unsafe_allow_html=True)
file1 = st.file_uploader("", type="xlsx", key="file1")

st.markdown('<label class="upload-label">📂 確定データファイルをアップロード</label>', unsafe_allow_html=True)
file2 = st.file_uploader("", type="xlsx", key="file2")

def build_map(df: pd.DataFrame) -> dict:
    m = {}
    for _, row in df.iterrows():
        code = row[0]
        if pd.isna(code):
            continue
        code = str(code).strip()
        name = row[1] if not pd.isna(row[1]) else ''
        qty = int(row[4]) if not pd.isna(row[4]) else 0
        after_sort = int(row[7]) if not pd.isna(row[7]) else 0
        m[code] = {'name': name, 'qty': qty, 'after_sort': after_sort}
    return m

def compute_diff(map1: dict, map2: dict) -> pd.DataFrame:
    rows = []
    for code, rec1 in map1.items():
        qty1 = rec1['qty']
        after1 = rec1['after_sort']
        if code in map2:
            rec2 = map2[code]
            qty2 = rec2.get('qty', 0)
            after = rec2.get('after_sort', after1)
        else:
            qty2 = 0
            after = after1
        diff = qty2 - qty1
        if diff != 0:
            rows.append({
                '商品コード': code,
                '商品名': rec1['name'],
                '暫定': qty1,
                '確定': qty2,
                '仕分後残': after,
                '増減数': f'+{diff}' if diff > 0 else str(diff)
            })
    for code, rec2 in map2.items():
        if code not in map1:
            qty2 = rec2['qty']
            after = rec2['after_sort']
            rows.append({
                '商品コード': code,
                '商品名': rec2['name'],
                '暫定': 0,
                '確定': qty2,
                '仕分後残': after,
                '増減数': f'+{qty2}'
            })
    df = pd.DataFrame(rows)
    df = df[~df['商品名'].str.startswith('■') & ~df['商品名'].str.endswith('◇')]
    df['__sort'] = df['商品名'].str[0].apply(lambda ch: 0 if ch == '■' else 1 if ch in ('□', '▢') else 2)
    df = df.sort_values(['__sort', '商品コード']).drop(columns='__sort')
    return df

def to_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = '差分結果'

    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill("solid", fgColor="CCFFCC")
    alt_fill = PatternFill("solid", fgColor="F2F2F2")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    headers = list(df.columns)
    ws.append(headers)

    col_widths = {
        'A': 9.0,
        'B': 38.25,
        'C': 9.0,
        'D': 9.0,
        'E': 9.0,
        'F': 9.0
    }

    for col_num, col in enumerate(headers, 1):
        col_letter = ws.cell(row=1, column=col_num).column_letter
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        if col_letter in col_widths:
            ws.column_dimensions[col_letter].width = col_widths[col_letter]

    ws.row_dimensions[1].height = 18.0  # ヘッダー行の高さ

    for idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.border = border
            if col_num in (1, 2):
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            if col_num == 5:  # E列（仕分後残）を太字に
                cell.font = Font(bold=True)
            if idx % 2 == 0:
                cell.fill = alt_fill
        ws.row_dimensions[ws.max_row].height = 14.5  # 各行の高さを14.5に設定

    # 最終行の1行下に日付付きコメントを挿入
    date_str = datetime.datetime.today().strftime('%-m/%-d(%a)').replace('Mon', '月').replace('Tue', '火').replace('Wed', '水').replace('Thu', '木').replace('Fri', '金').replace('Sat', '土').replace('Sun', '日')
    ws.cell(row=ws.max_row + 2, column=2, value=f"{date_str}仕分け分")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

if file1 and file2:
    df1 = pd.read_excel(file1, header=None).iloc[4:].reset_index(drop=True)
    df2 = pd.read_excel(file2, header=None).iloc[4:].reset_index(drop=True)
    diff_df = compute_diff(build_map(df1), build_map(df2))

    st.success("差分を抽出しました！")
    st.dataframe(diff_df, use_container_width=True)

    excel_data = to_excel(diff_df)
    st.download_button("📅 差分をExcelでダウンロード", excel_data, file_name=f"差分_{datetime.date.today()}.xlsx")
